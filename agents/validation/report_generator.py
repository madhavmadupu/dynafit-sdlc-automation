"""
agents/validation/report_generator.py
Phase 5 Excel Generator.
Transforms ValidatedFitmentBatch into fitment_matrix.xlsx.
"""

from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path

import openpyxl
import structlog
from openpyxl.styles import Alignment, Font, PatternFill

from core.config.settings import settings
from core.schemas.classification_result import ValidatedFitmentBatch
from core.schemas.enums import Verdict
from core.schemas.requirement_atom import RequirementAtom

log = structlog.get_logger()

# Cell coloring map for quick scanning
VERDICT_COLORS = {
    Verdict.FIT: "C6EFCE",  # Light Green
    Verdict.PARTIAL_FIT: "FFEB9C",  # Light Yellow
    Verdict.GAP: "FFC7CE",  # Light Red
}


def generate_excel_report(
    batch: ValidatedFitmentBatch,
    atoms: list[RequirementAtom],
) -> str:
    """
    Generate the final fitment_matrix.xlsx file.

    Columns:
    - Atom ID
    - Module
    - Priority
    - Requirement Text
    - Final Verdict (AI + Overrides combined)
    - Matched D365 Capability
    - Gap Description
    - Rationale (AI Chain-of-Thought or Consultant reason)
    - LLM Route Taken
    - Overridden By (if applicable)

    Args:
        batch: Final validation batch containing results and overrides.
        atoms: Original requirement atoms.

    Returns:
        Absolute path to the generated Excel file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fitment Matrix"

    # Header Definition
    headers = [
        "Atom ID",
        "Module",
        "Sub-module",
        "Priority",
        "Intent",
        "Country",
        "Completeness Score",
        "Source Ref",
        "Requirement Text",
        "Verdict",
        "Matched Capability",
        "Gap Description",
        "Configuration Needed",
        "Caveats",
        "Rationale",
        "Route Taken",
        "Sanity Flags",
        "Overridden By",
    ]

    # Write Headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    alignment = Alignment(vertical="top", wrap_text=True)

    ws.append(headers)
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment

    # Mapping atoms
    atom_map = {str(a.id): a for a in atoms}

    # Override Mapping
    override_map = {str(o.atom_id): o for o in batch.overrides}

    # Write Data
    for result in batch.results:
        atom_id = str(result.atom_id)
        atom = atom_map.get(atom_id)
        override = override_map.get(atom_id)

        overridden_by = override.reviewed_by if override else ""

        row = [
            atom_id,
            atom.module.value if atom else "",
            atom.sub_module if atom else "",
            atom.priority.value if atom else "",
            atom.intent.value if atom else "",
            atom.country if atom else "",
            f"{atom.completeness_score:.1f}" if atom else "",
            atom.source_ref if atom else "",
            atom.text if atom else "",
            result.verdict.value,
            result.matched_capability or "",
            result.gap_description or "",
            result.config_needed or "",
            "; ".join(result.caveats) if result.caveats else "",
            result.rationale,
            result.route_taken.value,
            "; ".join(result.sanity_flags) if result.sanity_flags else "",
            overridden_by,
        ]
        ws.append(row)

        # Color the verdict column
        last_row = ws.max_row
        verdict_cell = ws.cell(row=last_row, column=headers.index("Verdict") + 1)
        if result.verdict in VERDICT_COLORS:
            color = VERDICT_COLORS[result.verdict]
            verdict_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Column Widths
    column_widths = {
        "Atom ID": 36,
        "Module": 10,
        "Sub-module": 15,
        "Requirement Text": 60,
        "Verdict": 15,
        "Matched Capability": 30,
        "Gap Description": 40,
        "Rationale": 60,
        "Source Ref": 20,
    }

    for col_num, header in enumerate(headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        width = column_widths.get(header, 15)
        ws.column_dimensions[col_letter].width = width

    os.makedirs(settings.OUTPUT_DIR, exist_ok=True)
    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filepath = Path(settings.OUTPUT_DIR) / f"fitment_matrix_{batch.run_id}_{stamp}.xlsx"

    wb.save(str(filepath))
    log.info("report_generated", run_id=batch.run_id, path=str(filepath))

    return str(filepath)
